import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)

ZERO_THRESHOLD = 0.005

# Стандартные шрифты reportlab (Helvetica и т.п.) не содержат кириллических
# глифов — русский текст в PDF просто не отрисуется. Поэтому регистрируем
# шрифт DejaVu Sans, который лежит прямо в проекте (services/fonts/) —
# так экспорт работает одинаково на любой машине, без зависимости от того,
# какие шрифты установлены в системе пользователя.
_FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
_FONT_REGULAR = "DejaVuSans"
_FONT_BOLD = "DejaVuSans-Bold"

pdfmetrics.registerFont(TTFont(_FONT_REGULAR, os.path.join(_FONTS_DIR, "DejaVuSans.ttf")))
pdfmetrics.registerFont(TTFont(_FONT_BOLD, os.path.join(_FONTS_DIR, "DejaVuSans-Bold.ttf")))


def _balance_status(balance: float) -> str:
    """Единая формулировка статуса баланса — используется во всех трёх форматах,
    чтобы TXT/XLSX/PDF не разъезжались в формулировках."""
    if balance > ZERO_THRESHOLD:
        return f"ему должны {balance:.2f}"
    if balance < -ZERO_THRESHOLD:
        return f"должен {abs(balance):.2f}"
    return "в расчёте"


# --- TXT ---


def export_txt(report: dict) -> str:
    """Собирает текстовый отчёт: список трат, балансы, итоговые переводы.

    report — тот же словарь, что возвращает handlers.calculate.build_report_data.
    """
    event = report["event"]
    participants = report["participants"]
    expenses = report["expenses"]

    lines = [f"Итоги события «{event.title}»", "=" * 40, ""]

    lines.append("ТРАТЫ:")
    if expenses:
        for expense in expenses:
            desc = f" — {expense.description}" if expense.description else ""
            participants_names = ", ".join(s.participant.name for s in expense.shares)
            lines.append(
                f"  {expense.payer.name} заплатил {expense.amount:.2f}{desc} "
                f"(участвуют: {participants_names})"
            )
    else:
        lines.append("  Трат пока нет.")
    lines.append("")

    lines.append("БАЛАНСЫ:")
    if participants and report["balance"] is not None:
        for participant in participants:
            pid = participant.id
            status = _balance_status(report["balance"][pid])
            lines.append(
                f"  {participant.name}: заплатил {report['paid'][pid]:.2f}, "
                f"должен был {report['owed'][pid]:.2f} → {status}"
            )
    else:
        lines.append("  Недостаточно данных для расчёта.")
    lines.append("")

    lines.append("ПЕРЕВОДЫ:")
    transactions = report["transactions"]
    if transactions:
        name_by_id = report["name_by_id"]
        for debtor_id, creditor_id, amount in transactions:
            lines.append(f"  {name_by_id[debtor_id]} → {name_by_id[creditor_id]}: {amount:.2f}")
    else:
        lines.append("  Переводы не нужны.")

    return "\n".join(lines)


# --- XLSX ---


def export_xlsx(report: dict) -> bytes:
    """Собирает Excel-книгу с тремя листами: Траты, Балансы, Переводы."""
    event = report["event"]
    participants = report["participants"]
    expenses = report["expenses"]

    workbook = Workbook()
    header_font = Font(bold=True)

    # --- Лист "Траты" ---
    sheet_expenses = workbook.active
    sheet_expenses.title = "Траты"
    sheet_expenses.append(["Плательщик", "Сумма", "Описание", "Участники"])
    for cell in sheet_expenses[1]:
        cell.font = header_font

    for expense in expenses:
        participants_names = ", ".join(s.participant.name for s in expense.shares)
        sheet_expenses.append(
            [expense.payer.name, expense.amount, expense.description or "", participants_names]
        )

    _autosize_columns(sheet_expenses)

    # --- Лист "Балансы" ---
    sheet_balances = workbook.create_sheet("Балансы")
    sheet_balances.append(["Участник", "Заплатил", "Должен был", "Баланс", "Статус"])
    for cell in sheet_balances[1]:
        cell.font = header_font

    if participants and report["balance"] is not None:
        for participant in participants:
            pid = participant.id
            balance = report["balance"][pid]
            sheet_balances.append(
                [
                    participant.name,
                    report["paid"][pid],
                    report["owed"][pid],
                    balance,
                    _balance_status(balance),
                ]
            )

    _autosize_columns(sheet_balances)

    # --- Лист "Переводы" ---
    sheet_transfers = workbook.create_sheet("Переводы")
    sheet_transfers.append(["Кто переводит", "Кому", "Сумма"])
    for cell in sheet_transfers[1]:
        cell.font = header_font

    transactions = report["transactions"]
    if transactions:
        name_by_id = report["name_by_id"]
        for debtor_id, creditor_id, amount in transactions:
            sheet_transfers.append([name_by_id[debtor_id], name_by_id[creditor_id], amount])

    _autosize_columns(sheet_transfers)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _autosize_columns(sheet, min_width: int = 10, max_width: int = 40) -> None:
    """Простая автоширина колонок по самому длинному значению — openpyxl
    не делает это сам, а без неё длинные имена/описания обрезаются визуально."""
    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max(length + 2, min_width), max_width)
        # Выравнивание по левому краю для читаемости описаний/имён
        for cell in column_cells:
            cell.alignment = Alignment(horizontal="left")


# --- PDF ---


def export_pdf(report: dict) -> bytes:
    """Собирает PDF-отчёт: заголовок, таблица трат, таблица балансов, таблица переводов."""
    event = report["event"]
    participants = report["participants"]
    expenses = report["expenses"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleRu", parent=styles["Title"], fontName=_FONT_BOLD, fontSize=16, spaceAfter=12
    )
    heading_style = ParagraphStyle(
        "HeadingRu", parent=styles["Heading2"], fontName=_FONT_BOLD, fontSize=13,
        spaceBefore=16, spaceAfter=8,
    )
    normal_style = ParagraphStyle("NormalRu", parent=styles["Normal"], fontName=_FONT_REGULAR)

    elements = [Paragraph(f"Итоги события «{event.title}»", title_style)]

    # --- Таблица трат ---
    elements.append(Paragraph("Траты", heading_style))
    if expenses:
        expense_rows = [["Плательщик", "Сумма", "Описание", "Участники"]]
        for expense in expenses:
            participants_names = ", ".join(s.participant.name for s in expense.shares)
            expense_rows.append(
                [
                    expense.payer.name,
                    f"{expense.amount:.2f}",
                    expense.description or "",
                    participants_names,
                ]
            )
        elements.append(_styled_table(expense_rows, col_widths=[3 * cm, 2.5 * cm, 4 * cm, 7 * cm]))
    else:
        elements.append(Paragraph("Трат пока нет.", normal_style))

    # --- Таблица балансов ---
    elements.append(Paragraph("Балансы", heading_style))
    if participants and report["balance"] is not None:
        balance_rows = [["Участник", "Заплатил", "Должен был", "Статус"]]
        for participant in participants:
            pid = participant.id
            balance_rows.append(
                [
                    participant.name,
                    f"{report['paid'][pid]:.2f}",
                    f"{report['owed'][pid]:.2f}",
                    _balance_status(report["balance"][pid]),
                ]
            )
        elements.append(_styled_table(balance_rows, col_widths=[4 * cm, 3 * cm, 3 * cm, 6.5 * cm]))
    else:
        elements.append(Paragraph("Недостаточно данных для расчёта.", normal_style))

    # --- Таблица переводов ---
    elements.append(Paragraph("Переводы", heading_style))
    transactions = report["transactions"]
    if transactions:
        name_by_id = report["name_by_id"]
        transfer_rows = [["Кто переводит", "Кому", "Сумма"]]
        for debtor_id, creditor_id, amount in transactions:
            transfer_rows.append([name_by_id[debtor_id], name_by_id[creditor_id], f"{amount:.2f}"])
        elements.append(_styled_table(transfer_rows, col_widths=[5.5 * cm, 5.5 * cm, 5.5 * cm]))
    else:
        elements.append(Paragraph("Переводы не нужны.", normal_style))

    doc.build(elements)
    return buffer.getvalue()


def _styled_table(rows: list[list[str]], col_widths: list[float]) -> Table:
    """Общий стиль таблиц в PDF-отчёте — шапка выделена, границы тонкие."""
    table = Table(rows, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), _FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), _FONT_REGULAR),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
        )
    )
    return table